from __future__ import annotations

import json
import re
from collections import OrderedDict
from dataclasses import dataclass
from pathlib import Path
from typing import Any


REQUIRED_HEADERS = ["SKU ID", "SKU 名称", "原价", "计算价格", "库存", "SKU 图片"]
_SPLIT_RE = re.compile(r"\s*[、,，|/]\s*")


@dataclass(frozen=True)
class HuskySkuRow:
    sku_id: str
    sku_name: str
    original_price: float
    calculated_price: float
    stock: int
    image: str
    parts: list[str]


def _clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _as_float(value: Any, field: str, row_no: int) -> float:
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"第 {row_no} 行 {field} 不是有效数字: {value!r}") from exc


def _as_int(value: Any, field: str, row_no: int) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError) as exc:
        raise ValueError(f"第 {row_no} 行 {field} 不是有效整数: {value!r}") from exc


def _split_sku_name(name: str) -> list[str]:
    parts = [part.strip() for part in _SPLIT_RE.split(name) if part.strip()]
    return parts or [name.strip()]


def read_husky_xlsx(path: str | Path, sheet_name: str | None = None) -> list[HuskySkuRow]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("读取哈士奇 Excel 需要 openpyxl；请先执行 `pip install -e .`") from exc

    workbook = load_workbook(filename=Path(path), read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration as exc:
        raise ValueError("Excel 为空") from exc

    headers = [_clean(v) for v in header_row]
    missing = [name for name in REQUIRED_HEADERS if name not in headers]
    if missing:
        raise ValueError(f"缺少哈士奇导出列: {', '.join(missing)}")
    index = {name: headers.index(name) for name in REQUIRED_HEADERS}

    result: list[HuskySkuRow] = []
    for row_no, row in enumerate(rows, start=2):
        sku_id = _clean(row[index["SKU ID"]])
        sku_name = _clean(row[index["SKU 名称"]])
        if not sku_id and not sku_name:
            continue
        if not sku_id or not sku_name:
            raise ValueError(f"第 {row_no} 行 SKU ID / SKU 名称不能为空")
        result.append(
            HuskySkuRow(
                sku_id=sku_id,
                sku_name=sku_name,
                original_price=_as_float(row[index["原价"]], "原价", row_no),
                calculated_price=_as_float(row[index["计算价格"]], "计算价格", row_no),
                stock=_as_int(row[index["库存"]], "库存", row_no),
                image=_clean(row[index["SKU 图片"]]),
                parts=_split_sku_name(sku_name),
            )
        )
    if not result:
        raise ValueError("没有读取到 SKU 数据")
    return result


def normalize_husky_skus(rows: list[HuskySkuRow]) -> dict[str, Any]:
    max_parts = max(len(row.parts) for row in rows)
    columns: list[list[str]] = []
    for i in range(max_parts):
        values = OrderedDict()
        for row in rows:
            value = row.parts[i] if i < len(row.parts) else ""
            if value:
                values[value] = None
        columns.append(list(values))

    varying_axes: list[dict[str, Any]] = []
    common_specs: list[str] = []
    for i, values in enumerate(columns, start=1):
        if len(values) <= 1:
            if values:
                common_specs.append(values[0])
            continue
        varying_axes.append({"index": i, "name": f"规格{i}", "values": values})

    normalized_rows = []
    for row in rows:
        varying_values: list[str] = []
        for axis in varying_axes:
            idx = int(axis["index"]) - 1
            varying_values.append(row.parts[idx] if idx < len(row.parts) else "")
        normalized_rows.append(
            {
                "sku_id": row.sku_id,
                "sku_name": row.sku_name,
                "spec_values": varying_values,
                "original_price": row.original_price,
                "price": row.calculated_price,
                "stock": row.stock,
                "image": row.image,
            }
        )

    return {
        "source": "husky_xlsx",
        "sku_count": len(rows),
        "axes": varying_axes,
        "common_specs": common_specs,
        "rows": normalized_rows,
    }


def import_husky_xlsx(path: str | Path, sheet_name: str | None = None) -> dict[str, Any]:
    return normalize_husky_skus(read_husky_xlsx(path, sheet_name=sheet_name))


def save_normalized_json(data: dict[str, Any], output: str | Path) -> Path:
    target = Path(output)
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return target
